#!/usr/bin/env python3
"""Gera a planilha .xlsx mensal do organizador financeiro a partir dos CSVs
em financeiro/ (lancamentos.csv, gastos_fixos.csv, renda.csv, renda_variavel.csv).

Uso:
    python3 scripts/gerar_planilha_mensal.py --mes 9 --ano 2026 --out /tmp/saida.xlsx

Os CSVs são a única fonte de verdade. Este script nunca escreve neles —
só lê e gera o relatório do mês pedido.

renda.csv guarda só fontes com valor fixo conhecido de antemão (salário).
Renda de valor variável (reembolso, freelance) não cabe numa linha estática
com um único `valor` — entra em renda_variavel.csv como lançamento por mês,
no mesmo espírito de lancamentos.csv.
"""
import argparse
import csv
import re
import zipfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
FIN_DIR = REPO_ROOT / "financeiro"

CATEGORIAS = [
    "Moradia",
    "Contas e Utilidades",
    "Alimentação",
    "Transporte",
    "Saúde",
    "Educação",
    "Lazer e Assinaturas",
    "Vestuário",
    "Investimentos",
    "Dívidas e Financiamentos",
    "Outros",
]

MESES_PT = [
    "", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho",
    "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=11)
NEGATIVO_FILL = PatternFill("solid", fgColor="F8CBAD")
CURRENCY_FMT = '"R$" #,##0.00'
PCT_FMT = "0.0%"
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def ler_csv(nome):
    caminho = FIN_DIR / nome
    with caminho.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def to_float(valor, contexto):
    try:
        return float(valor)
    except (TypeError, ValueError):
        raise ValueError(
            f"Valor inválido ({valor!r}) em {contexto} — esperado número com ponto decimal."
        )


def validar_categoria(categoria, contexto):
    if categoria not in CATEGORIAS:
        raise ValueError(
            f"Categoria desconhecida ({categoria!r}) em {contexto}. "
            f"Categorias válidas: {', '.join(CATEGORIAS)}."
        )


def eh_ativo(valor):
    return (valor or "").strip().lower() == "sim"


def filtrar_mes(linhas, ano, mes, arquivo, exigir_categoria=False):
    saida = []
    for row in linhas:
        try:
            d = date.fromisoformat(row["data"].strip())
        except ValueError:
            raise ValueError(f"Data inválida em {arquivo}: {row!r}")
        if d.year == ano and d.month == mes:
            if exigir_categoria:
                validar_categoria(row["categoria"].strip(), f"{arquivo} ({row['data']})")
            saida.append(row)
    saida.sort(key=lambda r: r["data"])
    return saida


def estilo_header(ws, linha, n_colunas):
    for c in range(1, n_colunas + 1):
        cell = ws.cell(row=linha, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autofit(ws, larguras):
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def montar_renda(wb, renda):
    ws = wb.create_sheet("Renda")
    ws.append(["Descrição", "Valor", "Dia recebimento", "Tipo", "Ativo"])
    estilo_header(ws, 1, 5)
    for r in renda:
        ws.append([
            r["descricao"], to_float(r["valor"], "renda.csv"),
            r["dia_recebimento"], r["tipo"], r["ativo"],
        ])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=5):
        row[1].number_format = CURRENCY_FMT
        for c in row:
            c.border = BORDER
            c.font = Font(name=FONT_NAME)
    autofit(ws, [32, 14, 16, 12, 8])
    return len(renda)


def montar_gastos_fixos(wb, fixos):
    ws = wb.create_sheet("GastosFixos")
    ws.append(["Descrição", "Categoria", "Valor", "Dia vencimento", "Ativo"])
    estilo_header(ws, 1, 5)
    for r in fixos:
        validar_categoria(r["categoria"].strip(), f"gastos_fixos.csv ({r['descricao']})")
        ws.append([
            r["descricao"], r["categoria"], to_float(r["valor"], "gastos_fixos.csv"),
            r["dia_vencimento"], r["ativo"],
        ])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=5):
        row[2].number_format = CURRENCY_FMT
        for c in row:
            c.border = BORDER
            c.font = Font(name=FONT_NAME)
    autofit(ws, [32, 22, 14, 16, 8])
    return len(fixos)


def montar_renda_variavel(wb, renda_var_mes):
    ws = wb.create_sheet("RendaVariavel")
    ws.append(["Data", "Descrição", "Valor", "Observação"])
    estilo_header(ws, 1, 4)
    for r in renda_var_mes:
        ws.append([
            r["data"], r["descricao"], to_float(r["valor"], "renda_variavel.csv"),
            r.get("observacao", ""),
        ])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4):
        row[2].number_format = CURRENCY_FMT
        for c in row:
            c.border = BORDER
            c.font = Font(name=FONT_NAME)
    autofit(ws, [12, 32, 14, 30])
    return len(renda_var_mes)


def montar_orcamentos(wb, orcamentos):
    """Teto de orçamento por categoria — não é um gasto que sai da conta,
    é um limite de referência pra comparar com o gasto variável real.
    Devolve {categoria: (número da linha, valor)} só para os tetos ativos,
    usado pela aba Resumo para referenciar a célula certa sem reabrir o
    workbook."""
    ws = wb.create_sheet("Orcamentos")
    ws.append(["Categoria", "Teto mensal", "Ativo"])
    estilo_header(ws, 1, 3)
    orcamento_por_categoria = {}
    linha = 2
    for r in orcamentos:
        validar_categoria(r["categoria"].strip(), f"orcamentos.csv ({r['categoria']})")
        valor = to_float(r["teto_mensal"], "orcamentos.csv")
        ws.append([r["categoria"], valor, r["ativo"]])
        if eh_ativo(r["ativo"]):
            orcamento_por_categoria[r["categoria"]] = (linha, valor)
        linha += 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=3):
        row[1].number_format = CURRENCY_FMT
        for c in row:
            c.border = BORDER
            c.font = Font(name=FONT_NAME)
    autofit(ws, [30, 14, 8])
    return orcamento_por_categoria


def montar_lancamentos(wb, lancamentos_mes):
    ws = wb.create_sheet("Lancamentos")
    ws.append(["Data", "Categoria", "Descrição", "Valor", "Forma pagamento", "Observação"])
    estilo_header(ws, 1, 6)
    for r in lancamentos_mes:
        ws.append([
            r["data"], r["categoria"], r["descricao"],
            to_float(r["valor"], "lancamentos.csv"),
            r.get("forma_pagamento", ""), r.get("observacao", ""),
        ])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=6):
        row[3].number_format = CURRENCY_FMT
        for c in row:
            c.border = BORDER
            c.font = Font(name=FONT_NAME)
    autofit(ws, [12, 22, 36, 14, 18, 30])
    return len(lancamentos_mes)


def montar_resumo(
    wb, mes, ano, renda, fixos, lancamentos_mes, renda_var_mes,
    n_renda, n_fixos, n_lanc, n_renda_var, orcamento_por_categoria,
):
    """Cria a aba Resumo com fórmulas reais (SUMIFS/SUMIF), e devolve um
    dicionário {(aba, célula): valor} para injeção de cache — ver
    inject_formula_cache() sobre por que isso é necessário neste ambiente."""
    ws = wb.create_sheet("Resumo", 0)
    ws.sheet_view.showGridLines = False
    cache = {}

    ws["A1"] = f"Resumo financeiro — {MESES_PT[mes]}/{ano}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")

    renda_range = f"Renda!B2:B{max(2, 1 + n_renda)}"
    renda_ativo_range = f"Renda!E2:E{max(2, 1 + n_renda)}"
    fixos_range = f"GastosFixos!C2:C{max(2, 1 + n_fixos)}"
    fixos_ativo_range = f"GastosFixos!E2:E{max(2, 1 + n_fixos)}"
    lanc_range = f"Lancamentos!D2:D{max(2, 1 + n_lanc)}"
    lanc_cat_range = f"Lancamentos!B2:B{max(2, 1 + n_lanc)}"
    renda_var_range = f"RendaVariavel!C2:C{max(2, 1 + n_renda_var)}"

    total_renda_fixa = sum(to_float(r["valor"], "renda.csv") for r in renda if eh_ativo(r["ativo"]))
    total_renda_variavel = sum(to_float(r["valor"], "renda_variavel.csv") for r in renda_var_mes)
    total_renda = total_renda_fixa + total_renda_variavel
    total_fixos = sum(to_float(r["valor"], "gastos_fixos.csv") for r in fixos if eh_ativo(r["ativo"]))
    total_variavel = sum(to_float(r["valor"], "lancamentos.csv") for r in lancamentos_mes)
    total_saidas = total_fixos + total_variavel
    saldo = total_renda - total_saidas

    linhas = [
        (
            "Renda total (fixa ativa + variável do mês)",
            f"=SUMIFS({renda_range},{renda_ativo_range},\"sim\")+SUM({renda_var_range})",
            total_renda,
        ),
        ("Gastos fixos (ativos)", f"=SUMIFS({fixos_range},{fixos_ativo_range},\"sim\")", total_fixos),
        ("Gastos variáveis do mês", f"=SUM({lanc_range})", total_variavel),
        ("Total de saídas", "=B4+B5", total_saidas),
        ("Saldo do mês", "=B3-B6", saldo),
    ]
    linha = 3
    for rotulo, formula, valor in linhas:
        ws.cell(row=linha, column=1, value=rotulo).font = SECTION_FONT
        cell = ws.cell(row=linha, column=2, value=formula)
        cell.number_format = CURRENCY_FMT
        cell.border = BORDER
        cell.font = Font(name=FONT_NAME, bold=(rotulo == "Saldo do mês"))
        cache[("Resumo", cell.coordinate)] = round(valor, 2)
        linha += 1

    if saldo < 0:
        ws["B7"].fill = NEGATIVO_FILL
        ws["A8"] = "⚠ Gastos maiores que a renda ativa do mês."
        ws["A8"].font = Font(name=FONT_NAME, italic=True, color="C00000")

    linha = 10
    ws.cell(row=linha, column=1, value="Gastos variáveis por categoria").font = SECTION_FONT
    linha += 1
    for col, titulo in enumerate(
        ["Categoria", "Valor", "% do variável", "Teto mensal", "Saldo do teto"], start=1
    ):
        ws.cell(row=linha, column=col, value=titulo)
    estilo_header(ws, linha, 5)
    linha += 1

    categorias_por_valor = {}
    for r in lancamentos_mes:
        categorias_por_valor[r["categoria"]] = categorias_por_valor.get(r["categoria"], 0.0) + to_float(
            r["valor"], "lancamentos.csv"
        )

    for categoria in CATEGORIAS:
        valor_cat = round(categorias_por_valor.get(categoria, 0.0), 2)
        ws.cell(row=linha, column=1, value=categoria)
        cell_valor = ws.cell(
            row=linha, column=2,
            value=f'=SUMIF({lanc_cat_range},"{categoria}",{lanc_range})',
        )
        cell_valor.number_format = CURRENCY_FMT
        cache[("Resumo", cell_valor.coordinate)] = valor_cat

        pct = (valor_cat / total_variavel) if total_variavel else 0.0
        cell_pct = ws.cell(
            row=linha, column=3,
            value=f"=IFERROR(B{linha}/$B$5,0)",
        )
        cell_pct.number_format = PCT_FMT
        cache[("Resumo", cell_pct.coordinate)] = round(pct, 4)

        orcamento = orcamento_por_categoria.get(categoria)
        if orcamento is not None:
            linha_teto, teto_valor = orcamento
            cell_teto = ws.cell(row=linha, column=4, value=f"=Orcamentos!B{linha_teto}")
            cell_teto.number_format = CURRENCY_FMT
            cache[("Resumo", cell_teto.coordinate)] = teto_valor

            saldo_teto = round(teto_valor - valor_cat, 2)
            cell_saldo_teto = ws.cell(row=linha, column=5, value=f"=D{linha}-B{linha}")
            cell_saldo_teto.number_format = CURRENCY_FMT
            cache[("Resumo", cell_saldo_teto.coordinate)] = saldo_teto
            if saldo_teto < 0:
                cell_saldo_teto.fill = NEGATIVO_FILL

        for c in range(1, 6):
            ws.cell(row=linha, column=c).border = BORDER
            ws.cell(row=linha, column=c).font = Font(name=FONT_NAME)
        linha += 1

    autofit(ws, [30, 16, 14, 14, 14])
    return cache


def inject_formula_cache(xlsx_path, cache):
    """Grava no XML o valor calculado (<v>) ao lado de cada fórmula.

    Necessário porque o openpyxl nunca grava cache de fórmula, e o
    LibreOffice headless deste ambiente não recalcula de forma confiável
    (timeout confirmado mesmo com 90s) — sem isso, qualquer visualizador que
    não recalcule ao vivo (preview de celular, por exemplo) mostra 0 ou
    branco em vez do resultado. Mesmo problema e mesma correção já
    documentados em .claude/skills/contratacao-fornecedor/SKILL.md.
    """
    import xml.etree.ElementTree as ET

    xlsx_path = Path(xlsx_path)
    tmp_path = xlsx_path.with_suffix(".tmp.xlsx")

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        workbook_root = ET.fromstring(zin.read("xl/workbook.xml"))
        rels_root = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))

        ns_main = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
        ns_rid = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        ns_rel = "{http://schemas.openxmlformats.org/package/2006/relationships}"

        sheet_rids = {
            sheet.get("name"): sheet.get(ns_rid)
            for sheet in workbook_root.iter(f"{ns_main}sheet")
        }
        rid_targets = {
            rel.get("Id"): rel.get("Target")
            for rel in rels_root.iter(f"{ns_rel}Relationship")
        }

        sheet_files = {}
        for name, rid in sheet_rids.items():
            target = rid_targets[rid].lstrip("/")
            sheet_files[name] = target if target.startswith("xl/") else f"xl/{target}"

        por_arquivo = {}
        for (sheet_name, coord), value in cache.items():
            arquivo = sheet_files[sheet_name]
            por_arquivo.setdefault(arquivo, {})[coord] = value

        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                alvo = por_arquivo.get(item.filename)
                if alvo:
                    xml = data.decode("utf-8")
                    for coord, value in alvo.items():
                        # openpyxl já grava um <v/> vazio (ou <v></v>) depois da fórmula —
                        # substituir esse placeholder pelo valor calculado, não duplicar.
                        pattern = re.compile(
                            r'(<c r="%s"[^>]*>.*?<f>.*?</f>)(?:<v\s*/>|<v>.*?</v>)?(</c>)' % re.escape(coord),
                            re.DOTALL,
                        )
                        xml, n = pattern.subn(lambda m: f"{m.group(1)}<v>{value}</v>{m.group(2)}", xml, count=1)
                        if n == 0:
                            raise RuntimeError(f"Não encontrei a célula {coord} com fórmula em {item.filename}")
                    data = xml.encode("utf-8")
                zout.writestr(item, data)

    tmp_path.replace(xlsx_path)


def conferir(xlsx_path, cache):
    """Reabre com data_only=True e confere que o cache injetado bate."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True)
    for (sheet_name, coord), esperado in cache.items():
        lido = wb[sheet_name][coord].value
        if lido is None or round(float(lido), 2) != round(float(esperado), 2):
            raise RuntimeError(
                f"Conferência falhou: {sheet_name}!{coord} esperado {esperado}, lido {lido}"
            )


def gerar(mes, ano, saida):
    renda = ler_csv("renda.csv")
    fixos = ler_csv("gastos_fixos.csv")
    lancamentos = ler_csv("lancamentos.csv")
    renda_variavel = ler_csv("renda_variavel.csv")
    orcamentos = ler_csv("orcamentos.csv")
    lancamentos_mes = filtrar_mes(lancamentos, ano, mes, "lancamentos.csv", exigir_categoria=True)
    renda_var_mes = filtrar_mes(renda_variavel, ano, mes, "renda_variavel.csv")

    wb = Workbook()
    wb.remove(wb.active)

    n_renda = montar_renda(wb, renda)
    n_fixos = montar_gastos_fixos(wb, fixos)
    n_lanc = montar_lancamentos(wb, lancamentos_mes)
    n_renda_var = montar_renda_variavel(wb, renda_var_mes)
    orcamento_por_categoria = montar_orcamentos(wb, orcamentos)
    cache = montar_resumo(
        wb, mes, ano, renda, fixos, lancamentos_mes, renda_var_mes,
        n_renda, n_fixos, n_lanc, n_renda_var, orcamento_por_categoria,
    )

    saida = Path(saida)
    saida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(saida)

    inject_formula_cache(saida, cache)
    conferir(saida, cache)

    resumo = {
        "renda_total": cache[("Resumo", "B3")],
        "gastos_fixos": cache[("Resumo", "B4")],
        "gastos_variaveis": cache[("Resumo", "B5")],
        "saldo": cache[("Resumo", "B7")],
        "n_lancamentos": n_lanc,
    }
    return saida, resumo


def main():
    hoje = date.today()
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--mes", type=int, default=hoje.month, help="Mês (1-12)")
    ap.add_argument("--ano", type=int, default=hoje.year, help="Ano (AAAA)")
    ap.add_argument("--out", type=str, default=None, help="Caminho do .xlsx de saída")
    args = ap.parse_args()

    if not 1 <= args.mes <= 12:
        raise SystemExit(f"Mês inválido: {args.mes}")

    saida = args.out or f"/tmp/organizador_financeiro_{args.ano}-{args.mes:02d}.xlsx"
    caminho, resumo = gerar(args.mes, args.ano, saida)
    print(f"Planilha gerada e conferida: {caminho}")
    print(
        f"Renda ativa: R$ {resumo['renda_total']:.2f} | "
        f"Gastos fixos: R$ {resumo['gastos_fixos']:.2f} | "
        f"Gastos variáveis ({resumo['n_lancamentos']} lançamentos): R$ {resumo['gastos_variaveis']:.2f} | "
        f"Saldo: R$ {resumo['saldo']:.2f}"
    )


if __name__ == "__main__":
    main()
